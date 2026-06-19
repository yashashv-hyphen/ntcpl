#!/usr/bin/env python3
"""
Auto-fill Amazon category templates (.xlsm) from an initial product photo and product node.

Usage example:
  uv run --with openpyxl --with pillow python amazon_template_autofill.py \
    --template "LEASH_ANIMAL_COLLAR (1).xlsm" \
    --output "LEASH_FILLED.xlsm" \
    --initial-photo "/path/to/photo.jpg" \
    --product-node "Pet Supplies > Dogs > Collars, Harnesses & Leashes > Leashes > Basic Leashes (4771737031)" \
    --brand "Generic" \
    --product-name "Premium Leather Dog Leash"
"""

from __future__ import annotations

import argparse
import json
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import parse_qs
from urllib.request import urlopen

from huggingface_hub import InferenceClient
from openpyxl import load_workbook
from PIL import Image


@dataclass
class ColumnMeta:
    col_idx: int
    column_letter: str
    label: str
    attribute: str


@dataclass
class RequirementMeta:
    label: str
    requirement: str


def column_letter_from_index(idx: int) -> str:
    out = ""
    n = idx
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def parse_template_rows(template_ws) -> Tuple[int, int, int]:
    raw = str(template_ws["A1"].value or "").strip()
    if "settings=" in raw:
        raw = raw.split("settings=", 1)[1]
    params = parse_qs(raw, keep_blank_values=True)

    def parse_int(key: str, default: int) -> int:
        try:
            return int(params.get(key, [default])[0])
        except (TypeError, ValueError):
            return default

    label_row = parse_int("labelRow", 4)
    attribute_row = parse_int("attributeRow", 5)
    data_row = parse_int("dataRow", 7)
    return label_row, attribute_row, data_row


def read_column_metadata(template_ws, label_row: int, attribute_row: int) -> Dict[str, ColumnMeta]:
    meta: Dict[str, ColumnMeta] = {}
    for col_idx in range(1, template_ws.max_column + 1):
        label = str(template_ws.cell(label_row, col_idx).value or "").strip()
        attribute = str(template_ws.cell(attribute_row, col_idx).value or "").strip()
        if not attribute:
            continue
        meta[attribute] = ColumnMeta(
            col_idx=col_idx,
            column_letter=column_letter_from_index(col_idx),
            label=label,
            attribute=attribute,
        )
    return meta


def read_data_definitions(def_ws) -> Dict[str, RequirementMeta]:
    out: Dict[str, RequirementMeta] = {}
    for row in def_ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 6:
            continue
        attr        = str(row[1] or "").strip()  # col B
        label       = str(row[2] or "").strip()  # col C
        requirement = str(row[5] or "").strip()  # col F
        if attr:
            out[attr] = RequirementMeta(label=label, requirement=requirement)
    return out


def read_browse_nodes(browse_ws) -> List[Tuple[str, str]]:
    nodes: List[Tuple[str, str]] = []
    for row in browse_ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        node_id = str(row[0] or "").strip()  # col A
        path    = str(row[1] or "").strip()  # col B
        if node_id and path:
            nodes.append((node_id, path))
    return nodes


def read_recommended_node_to_product_type(valid_values_ws) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for row in valid_values_ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 3:
            continue
        key = str(row[1] or "").strip()  # col B
        val = str(row[2] or "").strip()  # col C
        if not key or not val:
            continue
        m = re.match(r"Recommended Browse Nodes - \[\s*(.+?)\s*\]$", key)
        if not m:
            continue
        product_type = m.group(1).strip()
        out[val] = product_type
    return out


def read_valid_product_types(valid_values_ws) -> List[str]:
    """Return all valid Product Type values from the Valid Values sheet.

    Rows whose column-B value matches 'Product Type - [...]' contribute all
    non-empty values from column C onwards to the result list.
    """
    types: List[str] = []
    seen: set = set()
    for row in valid_values_ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 2:
            continue
        key = str(row[1] or "").strip()  # col B
        if not re.match(r"Product Type\s*-\s*\[", key, re.I):
            continue
        for v_raw in row[2:]:  # col C+
            v = str(v_raw or "").strip()
            if v and v not in seen:
                seen.add(v)
                types.append(v)
    return types


def read_all_valid_values(valid_values_ws) -> Dict[str, List[str]]:
    """Read every constrained field from the Valid Values sheet.

    Each row has column B like "Color Map - [SHOES]" and column C+ with valid options.
    Returns {normalised_field_name: [valid_value, ...]} aggregated across all product types.
    Field names are lower-cased and stripped of the " - [ProductType]" suffix so callers
    can look up e.g. "color map", "size", "style", "condition type".

    Uses iter_rows() so it is efficient in both normal and read_only workbook modes.
    """
    out: Dict[str, List[str]] = {}
    seen_per_field: Dict[str, set] = {}
    _skip = re.compile(r"product type|recommended browse|product id type", re.I)
    for row in valid_values_ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 2:
            continue
        raw_key = str(row[1] or "").strip()  # col B
        if not raw_key or _skip.search(raw_key):
            continue
        # Strip " - [ProductType]" suffix to get the base field name
        field = re.sub(r"\s*-\s*\[.*?\]\s*$", "", raw_key).strip().lower()
        if not field:
            continue
        vals = out.setdefault(field, [])
        seen = seen_per_field.setdefault(field, set())
        for v_raw in row[2:]:  # col C+
            v = str(v_raw or "").strip()
            if v and v not in seen:
                seen.add(v)
                vals.append(v)
    return out


def match_product_node(user_node: str, nodes: List[Tuple[str, str]]) -> Tuple[str, str]:
    raw = user_node.strip()
    raw_lower = raw.lower()
    numeric = re.findall(r"\d+", raw)
    numeric = numeric[-1] if numeric else ""

    for node_id, path in nodes:
        full = f"{path} ({node_id})"
        if raw_lower == full.lower() or raw_lower == node_id.lower():
            return node_id, path
    for node_id, path in nodes:
        full = f"{path} ({node_id})"
        if raw_lower in full.lower() or full.lower() in raw_lower:
            return node_id, path
    if numeric:
        for node_id, path in nodes:
            if node_id == numeric:
                return node_id, path
    raise ValueError(f"Could not match product node '{user_node}' with Browse Data sheet.")


def load_image(path_or_url: str) -> Image.Image:
    if path_or_url.lower().startswith(("http://", "https://")):
        with urlopen(path_or_url) as resp:
            return Image.open(resp).convert("RGB")
    return Image.open(path_or_url).convert("RGB")


def _hf_token(explicit_token: Optional[str] = None) -> str:
    token = (
        explicit_token
        or os.environ.get("HF_TOKEN", "")
        or os.environ.get("HF_API_KEY", "")
        or os.environ.get("HUGGINGFACE_API_KEY", "")
    ).strip()
    if not token:
        raise ValueError("HF_TOKEN or HF_API_KEY is required for Hugging Face image generation.")
    return token


def _hf_client(explicit_token: Optional[str] = None) -> InferenceClient:
    return InferenceClient(token=_hf_token(explicit_token))


def _hf_image_model(explicit_model: Optional[str] = None) -> str:
    return (
        explicit_model
        or os.environ.get("HF_IMAGE_MODEL", "")
        or os.environ.get("HUGGINGFACE_IMAGE_MODEL", "")
        or "black-forest-labs/FLUX.1-schnell"
    ).strip()


def create_marketing_images(
    initial_photo: str,
    output_dir: Path,
    product_name: str,
    bullets: List[str],
    hf_token: Optional[str] = None,
    hf_image_model: Optional[str] = None,
) -> List[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    if not initial_photo.strip():
        raise ValueError("Initial photo is required for context-driven image prompts.")

    prompt_context = (
        f"Primary product image reference: {initial_photo.strip()}\n"
        f"Product name: {product_name.strip() or 'Generic Product'}\n"
        f"Top selling points: {', '.join([b for b in bullets[:3] if b.strip()]) or 'premium quality, durable use'}"
    )
    jobs = [
        (
            "amazon_main_white.jpg",
            "Create a pure white background Amazon main image with only the product centered, no text, no props.",
        ),
        (
            "amazon_feature_banner.jpg",
            "Create a clean feature-focused hero image composition for Amazon listing gallery, no logos or watermarks.",
        ),
        (
            "amazon_detail_zoom.jpg",
            "Create a close-up detail image highlighting texture/material quality for the product.",
        ),
        (
            "amazon_features.jpg",
            "Create an infographic-style features image with subtle design accents and no brand trademarks.",
        ),
        (
            "amazon_lifestyle.jpg",
            "Create a realistic lifestyle usage scene that naturally features the product.",
        ),
        (
            "amazon_dimensions.jpg",
            "Create a dimensions-focused technical style product image suitable for Amazon listing gallery.",
        ),
    ]
    client = _hf_client(hf_token)
    image_model = _hf_image_model(hf_image_model)

    generated: List[Path] = []
    for filename, prompt in jobs:
        out_path = output_dir / filename
        img = client.text_to_image(
            prompt=f"{prompt}\n{prompt_context}",
            model=image_model,
        )
        suffix = out_path.suffix.lower()
        fmt = "JPEG" if suffix in {".jpg", ".jpeg"} else "PNG"
        img.save(out_path, format=fmt)
        generated.append(out_path)

    return generated


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def first_attr_by_label(columns: Dict[str, ColumnMeta], label: str) -> Optional[str]:
    for attr, meta in columns.items():
        if normalize_text(meta.label).lower() == normalize_text(label).lower():
            return attr
    return None


def all_attrs_by_label(columns: Dict[str, ColumnMeta], label: str) -> List[str]:
    out: List[str] = []
    for attr, meta in columns.items():
        if normalize_text(meta.label).lower() == normalize_text(label).lower():
            out.append(attr)
    return out


def set_if_exists(ws, row: int, columns: Dict[str, ColumnMeta], attr: str, value: str) -> None:
    if attr in columns and value != "":
        ws.cell(row, columns[attr].col_idx).value = value


def get_if_exists(ws, row: int, columns: Dict[str, ColumnMeta], attr: str) -> str:
    if attr not in columns:
        return ""
    return str(ws.cell(row, columns[attr].col_idx).value or "").strip()


def ask_user_for_required(ws, data_row: int, columns: Dict[str, ColumnMeta], defs: Dict[str, RequirementMeta]) -> None:
    for attr, req in defs.items():
        if attr not in columns:
            continue
        requirement = req.requirement.lower()
        current = get_if_exists(ws, data_row, columns, attr)
        if current:
            continue

        label = req.label or columns[attr].label or attr
        if requirement == "required":
            while True:
                value = input(f"Required field '{label}' is missing. Enter value: ").strip()
                if value:
                    set_if_exists(ws, data_row, columns, attr, value)
                    break
        elif requirement == "conditionally required":
            yn = input(f"Conditionally required field '{label}' may be needed. Fill now? [y/N]: ").strip().lower()
            if yn == "y":
                value = input(f"Enter value for '{label}': ").strip()
                if value:
                    set_if_exists(ws, data_row, columns, attr, value)


def parse_extra_values(json_path: Optional[str]) -> Dict[str, str]:
    if not json_path:
        return {}
    content = Path(json_path).read_text(encoding="utf-8")
    raw = json.loads(content)
    if not isinstance(raw, dict):
        raise ValueError("--values-json must point to a JSON object.")
    return {str(k).strip(): str(v).strip() for k, v in raw.items() if str(v).strip()}


def apply_values(
    ws,
    data_row: int,
    columns: Dict[str, ColumnMeta],
    values: Dict[str, str],
) -> None:
    label_to_attrs: Dict[str, List[str]] = {}
    for attr, meta in columns.items():
        label_to_attrs.setdefault(normalize_text(meta.label).lower(), []).append(attr)

    for key, value in values.items():
        if key in columns:
            set_if_exists(ws, data_row, columns, key, value)
            continue
        attrs = label_to_attrs.get(normalize_text(key).lower(), [])
        if attrs:
            set_if_exists(ws, data_row, columns, attrs[0], value)


def build_defaults(product_name: str, brand: str) -> Tuple[str, str, List[str], str]:
    desc = (
        f"{product_name} by {brand} reads clearly from your main image—form, finish, and scale—so customers know what they are buying. "
        f"Everyday use cases come next: where and when someone reaches for this product, the routines it supports, and the problems it helps solve. "
        f"Practical usability ties it together—comfortable handling, easy care, straightforward setup or use—so the photo, title, "
        f"and story all point the same direction for the shopper."
    )
    bullets = [
        f"Photo-aligned: visible design, finish, and proportions of {product_name} so expectations match what arrives.",
        f"Use-case clarity: when and how customers use {product_name} in real life, not generic feature filler.",
        f"Usability for shoppers: ease of handling, daily comfort, or straightforward use suggested by your listing visuals.",
        f"Routine fit: how {product_name} fits the buyer's week—moments it improves and friction it removes.",
        f"Trust from transparency: what you see in the image supports {brand}'s practical benefits with less guesswork at checkout.",
    ]
    keywords = f"{product_name.lower()} {brand.lower()} premium durable marketplace"
    return desc, keywords, bullets, "New"


def main() -> None:
    parser = argparse.ArgumentParser(description="Auto-fill Amazon template from initial photo + product node.")
    parser.add_argument("--template", required=True, help="Path to input .xlsm template")
    parser.add_argument("--output", required=True, help="Path to output .xlsm file")
    parser.add_argument("--initial-photo", required=True, help="Product main photo (file path or URL)")
    parser.add_argument("--product-node", required=True, help="Browse node id/path, e.g. '(4771737031)'")
    parser.add_argument("--product-name", default="Generic Product", help="Product title base")
    parser.add_argument("--brand", default="Generic", help="Brand name")
    parser.add_argument("--values-json", help="Optional JSON mapping of {label_or_attribute: value}")
    parser.add_argument(
        "--generated-image-dir",
        default="generated_amazon_images",
        help="Directory for generated additional photos",
    )
    parser.add_argument(
        "--non-interactive",
        action="store_true",
        help="Do not prompt for missing required fields; fail instead.",
    )
    parser.add_argument(
        "--hf-token",
        default="",
        help="Hugging Face API token. If omitted, HF_TOKEN or HF_API_KEY environment variable is used.",
    )
    parser.add_argument(
        "--hf-image-model",
        default="",
        help="Optional image model override (defaults to HF_IMAGE_MODEL or HUGGINGFACE_IMAGE_MODEL).",
    )
    args = parser.parse_args()

    wb = load_workbook(args.template, keep_vba=True)
    if "Template" not in wb.sheetnames:
        raise ValueError("Template sheet not found.")
    template_ws = wb["Template"]
    data_def_ws = wb["Data Definitions"] if "Data Definitions" in wb.sheetnames else None
    browse_ws = wb["Browse Data"] if "Browse Data" in wb.sheetnames else None
    valid_values_ws = wb["Valid Values"] if "Valid Values" in wb.sheetnames else None

    label_row, attribute_row, data_row = parse_template_rows(template_ws)
    columns = read_column_metadata(template_ws, label_row, attribute_row)
    definitions = read_data_definitions(data_def_ws) if data_def_ws is not None else {}

    nodes = read_browse_nodes(browse_ws) if browse_ws is not None else []
    node_id, node_path = match_product_node(args.product_node, nodes) if nodes else ("", args.product_node)
    browse_full = f"{node_path} ({node_id})" if node_id else node_path

    node_to_pt = read_recommended_node_to_product_type(valid_values_ws) if valid_values_ws is not None else {}
    product_type = node_to_pt.get(browse_full)
    if not product_type and node_id:
        for k, v in node_to_pt.items():
            if f"({node_id})" in k:
                product_type = v
                break
    if not product_type:
        product_type = input("Could not infer Product Type. Enter Product Type exactly as template expects: ").strip()
        if not product_type:
            raise ValueError("Product Type is required.")

    sku = f"{re.sub(r'[^A-Za-z0-9]+', '-', args.product_name).strip('-')[:20]}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    listing_action = "Create or Replace (Full Update)"
    desc, keywords, bullets, condition_value = build_defaults(args.product_name, args.brand)

    direct_values: Dict[str, str] = {
        "SKU": sku,
        "Product Type": product_type,
        "Listing Action": listing_action,
        "Item Name": args.product_name,
        "Brand Name": args.brand,
        "Recommended Browse Nodes": browse_full,
        "Main Image URL": args.initial_photo,
        "Product Description": desc,
        "Bullet Point": bullets[0],
        "Generic Keyword": keywords,
        "Item Condition": condition_value,
    }

    apply_values(template_ws, data_row, columns, direct_values)

    bullet_attrs = all_attrs_by_label(columns, "Bullet Point")
    for i, attr in enumerate(bullet_attrs[:5]):
        if i < len(bullets):
            set_if_exists(template_ws, data_row, columns, attr, bullets[i])

    extra_values = parse_extra_values(args.values_json)
    apply_values(template_ws, data_row, columns, extra_values)

    gen_paths = create_marketing_images(
        initial_photo=args.initial_photo,
        output_dir=Path(args.generated_image_dir),
        product_name=args.product_name,
        bullets=bullets,
        hf_token=args.hf_token or None,
        hf_image_model=args.hf_image_model or None,
    )
    other_image_attrs = all_attrs_by_label(columns, "Other Image URL")
    for attr, img_path in zip(other_image_attrs, gen_paths[1:]):  # keep first image as main-style asset
        set_if_exists(template_ws, data_row, columns, attr, str(img_path.resolve()))

    main_image_attr = first_attr_by_label(columns, "Main Image URL")
    if main_image_attr and not get_if_exists(template_ws, data_row, columns, main_image_attr):
        set_if_exists(template_ws, data_row, columns, main_image_attr, str(gen_paths[0].resolve()))

    required_missing: List[str] = []
    for attr, req in definitions.items():
        if attr in columns and req.requirement.lower() == "required":
            if not get_if_exists(template_ws, data_row, columns, attr):
                required_missing.append(req.label or attr)

    if required_missing and args.non_interactive:
        missing = ", ".join(required_missing)
        raise ValueError(f"Missing required fields in non-interactive mode: {missing}")

    if not args.non_interactive:
        ask_user_for_required(template_ws, data_row, columns, definitions)

    wb.save(args.output)
    print(f"Saved: {args.output}")
    print(f"Generated images: {Path(args.generated_image_dir).resolve()}")


if __name__ == "__main__":
    main()

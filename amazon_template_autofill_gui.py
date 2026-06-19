#!/usr/bin/env python3
from __future__ import annotations

import os
import re
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from amazon_template_autofill import (
    ColumnMeta,
    RequirementMeta,
    all_attrs_by_label,
    build_defaults,
    create_marketing_images,
    match_product_node,
    parse_template_rows,
    read_browse_nodes,
    read_column_metadata,
    read_data_definitions,
    read_recommended_node_to_product_type,
)

APP_DIR = Path(__file__).resolve().parent


class AutoFillGUI:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Amazon Template AutoFill (Live Preview)")
        self.root.geometry("1400x860")

        self.template_path = tk.StringVar(value=str(APP_DIR / "LEASH_ANIMAL_COLLAR (1).xlsm"))
        self.output_path = tk.StringVar(value=str(APP_DIR / "LEASH_AUTOFILLED_GUI.xlsm"))
        self.initial_photo = tk.StringVar()
        self.product_node = tk.StringVar(value="4771737031")
        self.product_name = tk.StringVar(value="Generic Product")
        self.brand = tk.StringVar(value="Generic")
        self.product_type_manual = tk.StringVar(value="")
        self.generated_image_dir = tk.StringVar(value=str(APP_DIR / "generated_amazon_images"))
        self.hf_token = tk.StringVar(
            value=(
                os.environ.get("HF_TOKEN", "")
                or os.environ.get("HF_API_KEY", "")
                or os.environ.get("HUGGINGFACE_API_KEY", "")
            )
        )

        self.columns: Dict[str, ColumnMeta] = {}
        self.definitions: Dict[str, RequirementMeta] = {}
        self.required_attrs: List[str] = []
        self.nodes: List[Tuple[str, str]] = []
        self.node_to_product_type: Dict[str, str] = {}
        self.template_data_row: int = 7

        self.required_vars: Dict[str, tk.StringVar] = {}
        self.required_entry_widgets: Dict[str, tk.Entry] = {}

        self._build_layout()
        self._bind_live_updates()
        self.load_template()

    def _build_layout(self) -> None:
        top = ttk.Frame(self.root, padding=10)
        top.pack(fill=tk.X)

        def add_row(label: str, var: tk.StringVar, browse_cmd=None, show: Optional[str] = None):
            row = ttk.Frame(top)
            row.pack(fill=tk.X, pady=2)
            ttk.Label(row, text=label, width=21).pack(side=tk.LEFT)
            ttk.Entry(row, textvariable=var, show=show).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
            if browse_cmd:
                ttk.Button(row, text="Browse", command=browse_cmd).pack(side=tk.LEFT)

        add_row("Template (.xlsm)", self.template_path, self.browse_template)
        add_row("Output (.xlsm)", self.output_path, self.browse_output)
        add_row("Initial Photo Path/URL", self.initial_photo, self.browse_photo)
        add_row("Product Node", self.product_node)
        add_row("Product Name", self.product_name)
        add_row("Brand", self.brand)
        add_row("Product Type (optional)", self.product_type_manual)
        add_row("Generated Image Dir", self.generated_image_dir, self.browse_image_dir)
        add_row("HF API Token", self.hf_token, show="*")

        actions = ttk.Frame(top)
        actions.pack(fill=tk.X, pady=(8, 2))
        ttk.Button(actions, text="Reload Template", command=self.load_template).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(actions, text="Save Filled Excel", command=self.save_excel).pack(side=tk.LEFT)
        self.status_label = ttk.Label(actions, text="")
        self.status_label.pack(side=tk.LEFT, padx=12)

        middle = ttk.Panedwindow(self.root, orient=tk.HORIZONTAL)
        middle.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        left = ttk.Frame(middle)
        right = ttk.Frame(middle)
        middle.add(left, weight=2)
        middle.add(right, weight=3)

        ttk.Label(left, text="Required / Missing Fields").pack(anchor=tk.W)
        req_container = ttk.Frame(left)
        req_container.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self.req_canvas = tk.Canvas(req_container, borderwidth=0, highlightthickness=0)
        req_scroll = ttk.Scrollbar(req_container, orient=tk.VERTICAL, command=self.req_canvas.yview)
        self.req_inner = ttk.Frame(self.req_canvas)
        self.req_inner.bind(
            "<Configure>",
            lambda e: self.req_canvas.configure(scrollregion=self.req_canvas.bbox("all")),
        )
        self.req_canvas.create_window((0, 0), window=self.req_inner, anchor="nw")
        self.req_canvas.configure(yscrollcommand=req_scroll.set)
        self.req_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        req_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        ttk.Label(right, text="Live Template Row Preview (what will be written)").pack(anchor=tk.W)
        self.preview = ttk.Treeview(
            right,
            columns=("col", "label", "attr", "required", "value"),
            show="headings",
            height=30,
        )
        self.preview.heading("col", text="Col")
        self.preview.heading("label", text="Label")
        self.preview.heading("attr", text="Attribute")
        self.preview.heading("required", text="Required?")
        self.preview.heading("value", text="Value")
        self.preview.column("col", width=50, anchor=tk.CENTER)
        self.preview.column("label", width=210)
        self.preview.column("attr", width=430)
        self.preview.column("required", width=100, anchor=tk.CENTER)
        self.preview.column("value", width=520)
        self.preview.pack(fill=tk.BOTH, expand=True, pady=(4, 0))

    def _bind_live_updates(self) -> None:
        watched = [
            self.template_path,
            self.initial_photo,
            self.product_node,
            self.product_name,
            self.brand,
            self.product_type_manual,
            self.generated_image_dir,
        ]
        for var in watched:
            var.trace_add("write", lambda *_: self.refresh_preview())

    def browse_template(self) -> None:
        p = filedialog.askopenfilename(filetypes=[("Excel Macro Workbook", "*.xlsm"), ("Excel Workbook", "*.xlsx")])
        if p:
            self.template_path.set(p)
            if not self.output_path.get().strip():
                self.output_path.set(str(Path(p).with_name("AUTOFILLED.xlsm")))
            self.load_template()

    def browse_output(self) -> None:
        p = filedialog.asksaveasfilename(defaultextension=".xlsm", filetypes=[("Excel Macro Workbook", "*.xlsm")])
        if p:
            self.output_path.set(p)

    def browse_photo(self) -> None:
        p = filedialog.askopenfilename(
            filetypes=[("Images", "*.png *.jpg *.jpeg *.webp *.bmp *.tif *.tiff"), ("All files", "*.*")]
        )
        if p:
            self.initial_photo.set(p)

    def browse_image_dir(self) -> None:
        p = filedialog.askdirectory()
        if p:
            self.generated_image_dir.set(p)

    def load_template(self) -> None:
        try:
            wb = load_workbook(self.template_path.get(), keep_vba=True)
            ws = wb["Template"]
            dd = wb["Data Definitions"] if "Data Definitions" in wb.sheetnames else None
            bd = wb["Browse Data"] if "Browse Data" in wb.sheetnames else None
            vv = wb["Valid Values"] if "Valid Values" in wb.sheetnames else None

            label_row, attribute_row, data_row = parse_template_rows(ws)
            self.template_data_row = data_row
            self.columns = read_column_metadata(ws, label_row, attribute_row)
            self.definitions = read_data_definitions(dd) if dd is not None else {}
            self.nodes = read_browse_nodes(bd) if bd is not None else []
            self.node_to_product_type = read_recommended_node_to_product_type(vv) if vv is not None else {}

            self.required_attrs = [
                attr
                for attr, req in self.definitions.items()
                if attr in self.columns and req.requirement.strip().lower() == "required"
            ]
            self._render_required_inputs()
            self.refresh_preview()
            self.status_label.config(text="Template loaded.")
        except Exception as exc:
            self.status_label.config(text=f"Load failed: {exc}")
            messagebox.showerror("Template Load Error", str(exc))

    def _render_required_inputs(self) -> None:
        for child in self.req_inner.winfo_children():
            child.destroy()

        for attr in self.required_attrs:
            if attr not in self.required_vars:
                self.required_vars[attr] = tk.StringVar(value="")
                self.required_vars[attr].trace_add("write", lambda *_: self.refresh_preview())

            row = ttk.Frame(self.req_inner)
            row.pack(fill=tk.X, pady=2)
            label = self.definitions[attr].label or self.columns[attr].label or attr
            ttk.Label(row, text=label, width=32).pack(side=tk.LEFT)
            entry = ttk.Entry(row, textvariable=self.required_vars[attr])
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.required_entry_widgets[attr] = entry

    def _infer_product_type(self, browse_full: str, node_id: str) -> str:
        if self.product_type_manual.get().strip():
            return self.product_type_manual.get().strip()

        if browse_full in self.node_to_product_type:
            return self.node_to_product_type[browse_full]

        if node_id:
            for key, val in self.node_to_product_type.items():
                if f"({node_id})" in key:
                    return val
        return ""

    def _build_row_values(self) -> Tuple[Dict[str, str], List[str], str]:
        values: Dict[str, str] = {}
        node_id = ""
        browse_full = self.product_node.get().strip()

        if self.nodes and browse_full:
            try:
                node_id, node_path = match_product_node(browse_full, self.nodes)
                browse_full = f"{node_path} ({node_id})"
            except Exception:
                pass

        product_type = self._infer_product_type(browse_full, node_id)
        product_name = self.product_name.get().strip() or "Generic Product"
        brand = self.brand.get().strip() or "Generic"
        photo = self.initial_photo.get().strip()

        desc, keywords, bullets, condition_value = build_defaults(product_name, brand)
        sku = re.sub(r"[^A-Za-z0-9]+", "-", product_name).strip("-")[:20] or "SKU"

        by_label = {
            "SKU": sku,
            "Product Type": product_type,
            "Listing Action": "Create or Replace (Full Update)",
            "Item Name": product_name,
            "Brand Name": brand,
            "Recommended Browse Nodes": browse_full,
            "Main Image URL": photo,
            "Product Description": desc,
            "Generic Keyword": keywords,
            "Item Condition": condition_value,
        }
        label_to_attr: Dict[str, str] = {}
        for attr, meta in self.columns.items():
            label_to_attr.setdefault(meta.label.strip().lower(), attr)
        for lbl, val in by_label.items():
            attr = label_to_attr.get(lbl.lower())
            if attr and val:
                values[attr] = val

        bullet_attrs = all_attrs_by_label(self.columns, "Bullet Point")
        for i, attr in enumerate(bullet_attrs[:5]):
            if i < len(bullets):
                values[attr] = bullets[i]

        for attr in self.required_attrs:
            manual = self.required_vars.get(attr, tk.StringVar(value="")).get().strip()
            if manual:
                values[attr] = manual

        return values, bullets, browse_full

    def refresh_preview(self) -> None:
        values, _, _ = self._build_row_values()
        for item in self.preview.get_children():
            self.preview.delete(item)

        ordered = sorted(self.columns.values(), key=lambda c: c.col_idx)
        for meta in ordered:
            req = self.definitions.get(meta.attribute)
            req_txt = req.requirement if req else ""
            is_required = req_txt.strip().lower() == "required"
            v = values.get(meta.attribute, "")
            if is_required or v:
                self.preview.insert(
                    "",
                    tk.END,
                    values=(meta.column_letter, meta.label, meta.attribute, req_txt, v),
                )

    def _ask_for_missing_required(self, row_values: Dict[str, str]) -> bool:
        for attr in self.required_attrs:
            if row_values.get(attr, "").strip():
                continue
            label = self.definitions[attr].label or self.columns[attr].label or attr
            value = simpledialog.askstring("Required Field", f"Enter required field value:\n{label}")
            if not value:
                return False
            row_values[attr] = value.strip()
            if attr in self.required_vars:
                self.required_vars[attr].set(value.strip())
        return True

    def save_excel(self) -> None:
        try:
            if not self.template_path.get().strip():
                raise ValueError("Template path is required.")
            if not self.output_path.get().strip():
                raise ValueError("Output path is required.")
            if not self.initial_photo.get().strip():
                raise ValueError("Initial photo is required.")
            if not self.product_node.get().strip():
                raise ValueError("Product node is required.")

            row_values, bullets, _ = self._build_row_values()
            if not self._ask_for_missing_required(row_values):
                self.status_label.config(text="Save cancelled (required value missing).")
                return

            wb = load_workbook(self.template_path.get(), keep_vba=True)
            ws = wb["Template"]
            label_row, attribute_row, data_row = parse_template_rows(ws)
            columns = read_column_metadata(ws, label_row, attribute_row)

            for attr, value in row_values.items():
                if attr in columns and value:
                    ws.cell(data_row, columns[attr].col_idx).value = value

            generated = create_marketing_images(
                initial_photo=self.initial_photo.get().strip(),
                output_dir=Path(self.generated_image_dir.get().strip() or "generated_amazon_images"),
                product_name=self.product_name.get().strip() or "Generic Product",
                bullets=bullets,
                hf_token=self.hf_token.get().strip() or None,
            )

            main_attrs = all_attrs_by_label(columns, "Main Image URL")
            if main_attrs:
                ws.cell(data_row, columns[main_attrs[0]].col_idx).value = self.initial_photo.get().strip()

            other_attrs = all_attrs_by_label(columns, "Other Image URL")
            for attr, image_path in zip(other_attrs, generated[1:]):
                ws.cell(data_row, columns[attr].col_idx).value = str(image_path.resolve())

            wb.save(self.output_path.get().strip())
            self.refresh_preview()
            self.status_label.config(text=f"Saved: {self.output_path.get().strip()}")
            messagebox.showinfo(
                "Done",
                f"Excel saved:\n{self.output_path.get().strip()}\n\nGenerated images:\n{Path(self.generated_image_dir.get()).resolve()}",
            )
        except Exception as exc:
            self.status_label.config(text=f"Save failed: {exc}")
            messagebox.showerror("Save Error", str(exc))


def main() -> None:
    root = tk.Tk()
    style = ttk.Style()
    if "clam" in style.theme_names():
        style.theme_use("clam")
    AutoFillGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

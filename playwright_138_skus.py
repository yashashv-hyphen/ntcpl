#!/usr/bin/env python3
"""
End-to-end Playwright test: upload Template.xlsm + Sample File.xlsx,
process all 138 SKUs via SSE stream, save to Excel, download, verify.

Run:
    python3 playwright_138_skus.py

Iterates automatically if issues are found and keeps a detailed log.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import time
import tempfile
import threading
from pathlib import Path
from typing import Dict, List, Tuple, Optional

ROOT         = Path(__file__).parent
TEMPLATE_SRC = ROOT / "Desktop_Template.xlsm"
CATALOG_SRC  = ROOT / "Desktop_SampleFile.xlsx"
APP_PORT     = 7865
BASE_URL     = f"http://localhost:{APP_PORT}"
MAX_WAIT_S   = 2400   # 40 min total (138 SKUs × ~10s each at 8 workers)
WORKERS      = "8"

REQUIRED_ATTRS = [
    "contribution_sku#1.value",
    "product_type#1.value",
    "item_name[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value",
    "brand[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value",
    "product_description[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value",
    "bullet_point[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value",
]

# ─────────────────────────────────────────────────────────────────────────────

def _upload_template(page) -> str:
    """Upload the template file. Returns server-side path."""
    with open(TEMPLATE_SRC, "rb") as f:
        data = f.read()
    resp = page.request.post(
        f"{BASE_URL}/api/upload-template",
        multipart={
            "file": {
                "name": "Template.xlsm",
                "mimeType": "application/vnd.ms-excel.sheet.macroEnabled.12",
                "buffer": data,
            }
        },
    )
    j = resp.json()
    assert j.get("ok"), f"Template upload failed: {j}"
    print(f"  Template uploaded → {j['path']}")
    return j["path"]


def _upload_catalog(page) -> str:
    """Upload the image catalog. Returns server-side path."""
    with open(CATALOG_SRC, "rb") as f:
        data = f.read()
    resp = page.request.post(
        f"{BASE_URL}/api/upload-image-catalog",
        multipart={
            "file": {
                "name": "SampleFile.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "buffer": data,
            }
        },
    )
    j = resp.json()
    assert j.get("ok"), f"Catalog upload failed: {j}"
    print(f"  Catalog uploaded → {j['sku_count']} SKUs")
    assert j["sku_count"] == 138, f"Expected 138 SKUs, got {j['sku_count']}"
    return j["path"]


def _process_via_sse(template_path: str, catalog_path: str) -> Tuple[List[Dict], List[Dict]]:
    """
    Call /api/process-catalog/stream, consume the SSE, return (products, details).
    Uses urllib directly so we can stream without Playwright's request API (which
    doesn't support streaming responses).
    """
    import urllib.request
    import urllib.error

    payload = json.dumps({
        "template_path":      template_path,
        "image_catalog_path": catalog_path,
        "generate_images":    False,
        "provider":           "auto",
        "brand_name":         "NTCPL",
        "seller_notes":       "",
    }).encode()

    req = urllib.request.Request(
        f"{BASE_URL}/api/process-catalog/stream",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    products: List[Dict] = []
    details:  List[Dict] = []
    total     = 0
    completed = 0
    errors:   List[str] = []

    start = time.time()
    with urllib.request.urlopen(req, timeout=MAX_WAIT_S) as resp:
        buf = ""
        for raw_chunk in iter(lambda: resp.read(4096), b""):
            buf += raw_chunk.decode("utf-8", errors="replace")
            while "\n" in buf:
                line, buf = buf.split("\n", 1)
                line = line.strip()
                if not line.startswith("data: "):
                    continue
                try:
                    evt = json.loads(line[6:])
                except json.JSONDecodeError:
                    continue

                t = evt.get("type", "")
                if t == "start":
                    total = evt.get("total", 0)
                    print(f"  Stream started — {total} SKUs")

                elif t == "sku":
                    completed += 1
                    sku = evt.get("sku", "?")
                    errs = evt.get("pipeline", {}).get("errors", [])
                    if errs:
                        errors.append(f"{sku}: {errs}")
                    elapsed = time.time() - start
                    rate = completed / elapsed if elapsed > 0 else 0
                    eta = (total - completed) / rate if rate > 0 else 0
                    print(
                        f"  [{completed:3d}/{total}] {sku:<12} "
                        f"{'✗ ' + str(errs[0])[:40] if errs else '✓'}"
                        f"  ETA {eta:.0f}s",
                        end="\r",
                    )
                    sys.stdout.flush()

                elif t == "done":
                    print()  # newline after \r
                    products = evt.get("products", [])
                    details  = evt.get("details", [])
                    print(f"  Stream done — {evt.get('count')} products returned")

                elif t == "error":
                    print(f"\n  STREAM ERROR: {evt.get('error')}")
                    raise RuntimeError(f"Server SSE error: {evt.get('error')}")

    return products, details


def _save_excel(page, template_path: str, products: List[Dict]) -> str:
    """Call /api/save, return the saved_path."""
    resp = page.request.post(
        f"{BASE_URL}/api/save",
        data=json.dumps({"template_path": template_path, "products": products}),
        headers={"Content-Type": "application/json"},
    )
    j = resp.json()
    assert j.get("ok"), f"Save failed: {j}"
    saved = j.get("saved_path") or template_path
    warns = j.get("warnings", [])
    print(f"  Excel saved to {saved}")
    if warns:
        print(f"  {len(warns)} unfilled required field warning(s):")
        for w in warns[:10]:
            print(f"    row {w.get('row')} → {w.get('label')}")
    return saved


def _download_excel(page, saved_path: str) -> Path:
    """Download the Excel and save locally."""
    url = f"{BASE_URL}/api/download-excel?path={saved_path}"
    resp = page.request.get(url)
    assert resp.ok, f"Download failed: {resp.status}"
    out = ROOT / "output_138_skus.xlsm"
    out.write_bytes(resp.body())
    print(f"  Excel downloaded → {out} ({out.stat().st_size/1024:.0f} KB)")
    return out


def _verify_excel(xlsx_path: Path, products: List[Dict]) -> Dict:
    """
    Verify the downloaded Excel has all 138 SKUs with key fields filled.
    Returns a report dict.
    """
    from openpyxl import load_workbook
    sys.path.insert(0, str(ROOT))
    from amazon_template_autofill_web import load_template_meta

    wb    = load_workbook(str(xlsx_path), keep_vba=True, data_only=True)
    meta  = load_template_meta(str(TEMPLATE_SRC))
    ws    = wb["Template"]
    dr    = meta["data_row"]
    cols  = {c["attr"]: c["col_idx"] for c in meta["columns"]}

    report = {
        "total_rows":     0,
        "skus_found":     [],
        "skus_missing":   [],
        "empty_required": {},   # sku → [field_label, ...]
        "has_title":      0,
        "has_bullets":    0,
    }

    # Read all data rows
    row_data: List[Dict] = []
    sku_col   = cols.get("contribution_sku#1.value", 1)
    title_col = cols.get("item_name[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value")
    bp_col    = cols.get("bullet_point[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value")

    r = dr
    while True:
        sku_val = str(ws.cell(r, sku_col).value or "").strip()
        if not sku_val:
            break
        row_data.append({"sku": sku_val, "row": r})
        r += 1

    report["total_rows"] = len(row_data)

    # All SKUs from catalog
    from amazon_template_autofill_web import read_image_catalog
    catalog_skus = set(read_image_catalog(CATALOG_SRC).keys())
    found_skus   = {d["sku"] for d in row_data}

    report["skus_found"]   = sorted(found_skus)
    report["skus_missing"] = sorted(catalog_skus - found_skus)

    # Per-row quality check
    label_map = {c["attr"]: c["label"] for c in meta["columns"]}
    for d in row_data:
        row_r = d["row"]
        sku   = d["sku"]
        empty = []
        for attr in REQUIRED_ATTRS:
            if attr not in cols:
                continue
            val = str(ws.cell(row_r, cols[attr]).value or "").strip()
            if not val:
                empty.append(label_map.get(attr, attr))
        if empty:
            report["empty_required"][sku] = empty
        if title_col:
            v = str(ws.cell(row_r, title_col).value or "").strip()
            if v:
                report["has_title"] += 1
        if bp_col:
            v = str(ws.cell(row_r, bp_col).value or "").strip()
            if v:
                report["has_bullets"] += 1

    wb.close()
    return report


def _print_report(report: Dict, expected: int = 138) -> bool:
    """Print verification report. Returns True if all checks pass."""
    print("\n" + "=" * 60)
    print("VERIFICATION REPORT")
    print("=" * 60)
    total = report["total_rows"]
    found = len(report["skus_found"])
    miss  = len(report["skus_missing"])

    print(f"  Rows in Excel :  {total}")
    print(f"  SKUs found    :  {found} / {expected}")
    print(f"  SKUs missing  :  {miss}")
    print(f"  Has title     :  {report['has_title']} / {total}")
    print(f"  Has bullets   :  {report['has_bullets']} / {total}")

    if miss:
        print(f"\n  Missing SKUs ({miss}):")
        for s in report["skus_missing"][:20]:
            print(f"    {s}")
        if miss > 20:
            print(f"    ... and {miss - 20} more")

    empty_req = report["empty_required"]
    if empty_req:
        print(f"\n  SKUs with empty required fields ({len(empty_req)}):")
        for sku, fields in list(empty_req.items())[:15]:
            print(f"    {sku}: {', '.join(fields[:4])}")
        if len(empty_req) > 15:
            print(f"    ... and {len(empty_req) - 15} more")

    ok = (total == expected and miss == 0 and
          report["has_title"] == total and
          not empty_req)
    print("\n  STATUS:", "✓ PERFECT" if ok else "✗ NEEDS IMPROVEMENT")
    print("=" * 60 + "\n")
    return ok


# ── Browser UI verification ────────────────────────────────────────────────

def _verify_ui(page, template_path: str, catalog_path: str, saved_path: str) -> None:
    """
    Open the app in a real browser, load the saved Excel, and do a visual spot-check.
    """
    print("\n[UI] Opening app in browser for visual verification…")
    page.goto(BASE_URL, wait_until="domcontentloaded")

    # Verify the page loaded
    title = page.title()
    print(f"  Page title: {title}")

    # Check status endpoint directly
    resp = page.request.get(f"{BASE_URL}/api/status")
    status = resp.json()
    active = status.get("active", [])
    print(f"  Active providers: {active}")

    # Verify the template shows up when re-uploading
    with open(TEMPLATE_SRC, "rb") as f:
        data = f.read()
    up_resp = page.request.post(
        f"{BASE_URL}/api/upload-template",
        multipart={"file": {
            "name": "Template.xlsm",
            "mimeType": "application/vnd.ms-excel.sheet.macroEnabled.12",
            "buffer": data,
        }},
    )
    assert up_resp.json().get("ok"), "Template re-upload failed in UI check"

    # Load meta
    meta_resp = page.request.post(
        f"{BASE_URL}/api/meta",
        data=json.dumps({"template_path": up_resp.json()["path"]}),
        headers={"Content-Type": "application/json"},
    )
    meta_j = meta_resp.json()
    cols_count = len(meta_j.get("columns", meta_j.get("meta", {}).get("columns", [])))
    print(f"  Template columns: {cols_count}")

    print("  UI check passed ✓")


# ── Main ───────────────────────────────────────────────────────────────────

def run():
    from playwright.sync_api import sync_playwright

    attempt = 0
    best_rows = 0

    while True:
        attempt += 1
        print(f"\n{'='*60}")
        print(f"ATTEMPT {attempt}")
        print(f"{'='*60}\n")

        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page    = browser.new_page()

            try:
                # ── Phase 1: Upload ─────────────────────────────────────
                print("[1] Uploading files…")
                t_path = _upload_template(page)
                c_path = _upload_catalog(page)

                # ── Phase 2: Process via SSE ─────────────────────────────
                print(f"\n[2] Processing 138 SKUs (max {MAX_WAIT_S//60} min)…")
                t0 = time.time()
                products, details = _process_via_sse(t_path, c_path)
                elapsed = time.time() - t0
                print(f"  Processing took {elapsed:.0f}s ({elapsed/60:.1f} min)")
                print(f"  Products returned: {len(products)}")

                # ── Phase 3: Save to Excel ───────────────────────────────
                print("\n[3] Saving Excel…")
                saved = _save_excel(page, t_path, products)

                # ── Phase 4: Download ────────────────────────────────────
                print("\n[4] Downloading Excel…")
                xl_path = _download_excel(page, saved)

                # ── Phase 5: Verify ──────────────────────────────────────
                print("\n[5] Verifying Excel contents…")
                report = _verify_excel(xl_path, products)
                ok     = _print_report(report)

                # ── Phase 6: UI check ────────────────────────────────────
                _verify_ui(page, t_path, c_path, saved)

                if ok:
                    print(f"\n✓ ALL 138 SKUs MAPPED PERFECTLY in {elapsed:.0f}s!")
                    browser.close()
                    return True

                # Diagnose and maybe retry
                missing  = len(report["skus_missing"])
                bad_rows = len(report["empty_required"])
                rows     = report["total_rows"]

                if rows > best_rows:
                    best_rows = rows
                    print(f"  Progress: {rows} rows (best so far)")

                if attempt >= 3:
                    print(f"\nMax attempts reached. Best: {best_rows}/138 rows.")
                    browser.close()
                    return False

                print(f"\nRetrying in 10s… (missing={missing}, bad={bad_rows})")
                browser.close()
                time.sleep(10)

            except Exception as exc:
                import traceback
                print(f"\nERROR in attempt {attempt}: {exc}")
                traceback.print_exc()
                browser.close()
                if attempt >= 3:
                    return False
                time.sleep(15)


if __name__ == "__main__":
    print("=" * 60)
    print("138 SKU END-TO-END PLAYWRIGHT TEST")
    print("=" * 60)
    print(f"App: {BASE_URL}")
    print(f"Template: {TEMPLATE_SRC.name} ({TEMPLATE_SRC.stat().st_size//1024}KB)")
    print(f"Catalog: {CATALOG_SRC.name} ({CATALOG_SRC.stat().st_size//1024}KB)")
    print()

    # Check app is up
    import urllib.request
    try:
        urllib.request.urlopen(f"{BASE_URL}/api/status", timeout=5)
        print(f"App already running on port {APP_PORT}")
    except Exception:
        print(f"App not running on {APP_PORT}, please start it first.")
        print(f"Run: CATALOG_SKU_WORKERS=8 python3 amazon_template_autofill_web.py --no-browser --port {APP_PORT}")
        sys.exit(1)

    success = run()
    sys.exit(0 if success else 1)

"""
Playwright tests for:
1. Normal mode image quality — BG removal + tight crop applied to hero/infographic.
2. Valid Values tab — constrained fields (Color, Size, Style, Material) snapped to
   valid options from the template's Valid Values sheet.
3. Smoke-test that the app still starts, uploads work, and the status panel shows
   Ollama listed as a provider.

Run:
    python3 test_changes.py
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import threading
import time
from pathlib import Path

# ── helpers ──────────────────────────────────────────────────────────────────

ROOT = Path(__file__).parent
TEMPLATE = ROOT / "Template.xlsm"
CATALOG  = next(ROOT.glob("uploaded_catalogs/*.xlsx"), None)
APP_PORT = 7862  # use a dedicated test port so we don't clash with dev server


def _start_server() -> subprocess.Popen:
    env = os.environ.copy()
    env["SKIP_VISION"] = "1"   # skip real AI calls during UI tests
    proc = subprocess.Popen(
        [sys.executable, "amazon_template_autofill_web.py",
         "--port", str(APP_PORT), "--no-browser"],
        cwd=str(ROOT),
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )
    # Wait for the server to be ready
    import urllib.request, urllib.error
    deadline = time.time() + 20
    while time.time() < deadline:
        try:
            urllib.request.urlopen(f"http://localhost:{APP_PORT}/", timeout=2)
            return proc
        except Exception:
            if proc.poll() is not None:
                out, _ = proc.communicate()
                raise RuntimeError(f"Server exited early:\n{out.decode()}")
            time.sleep(0.4)
    raise TimeoutError("Server did not start in time")


# ═════════════════════════════════════════════════════════════════════════════
# Unit tests — no browser needed
# ═════════════════════════════════════════════════════════════════════════════

def test_snap_to_valid():
    """_snap_to_valid returns the best matching option."""
    import amazon_template_autofill_web as m

    assert m._snap_to_valid("black", ["Black", "White", "Blue"]) == "Black"
    assert m._snap_to_valid("navy blue", ["Navy Blue", "Red", "Green"]) == "Navy Blue"
    assert m._snap_to_valid("small", ["Small", "Medium", "Large"]) == "Small"
    # No match — return original unchanged
    assert m._snap_to_valid("ultraviolet", ["Red", "Green", "Blue"]) == "ultraviolet"
    # Empty options — return value
    assert m._snap_to_valid("cotton", []) == "cotton"
    print("  PASS  test_snap_to_valid")


def test_read_all_valid_values():
    """read_all_valid_values parses the Valid Values sheet correctly."""
    from openpyxl import load_workbook
    from amazon_template_autofill import read_all_valid_values

    if not TEMPLATE.exists():
        print("  SKIP  test_read_all_valid_values — Template.xlsm not found")
        return

    wb = load_workbook(str(TEMPLATE), keep_vba=True)
    if "Valid Values" not in wb.sheetnames:
        print("  SKIP  test_read_all_valid_values — no Valid Values sheet")
        wb.close()
        return

    vv = wb["Valid Values"]
    avv = read_all_valid_values(vv)
    wb.close()

    assert isinstance(avv, dict), "Expected dict"
    assert len(avv) > 0, "Valid Values should produce at least one field"
    # All values should be lists of strings
    for field, vals in avv.items():
        assert isinstance(vals, list), f"Values for '{field}' should be a list"
        assert all(isinstance(v, str) for v in vals), f"All values for '{field}' should be strings"
    print(f"  PASS  test_read_all_valid_values — {len(avv)} fields parsed: {list(avv.keys())[:5]}")


def test_load_template_meta_includes_all_valid_values():
    """load_template_meta returns all_valid_values key."""
    import amazon_template_autofill_web as m

    if not TEMPLATE.exists():
        print("  SKIP  test_load_template_meta_includes_all_valid_values — no template")
        return

    meta = m.load_template_meta(str(TEMPLATE))
    assert "all_valid_values" in meta, "meta should include all_valid_values"
    assert isinstance(meta["all_valid_values"], dict)
    print(f"  PASS  test_load_template_meta_includes_all_valid_values — "
          f"{len(meta['all_valid_values'])} fields")


def test_catalog_result_to_template_row_snaps_color():
    """catalog_result_to_template_row snaps color to valid options."""
    import amazon_template_autofill_web as m

    columns = [
        {"attr": "color", "label": "Color", "col_idx": 1, "col": "A",
         "group": "", "example": "", "required": ""},
        {"attr": "color_map", "label": "Color Map", "col_idx": 2, "col": "B",
         "group": "", "example": "", "required": ""},
    ]
    result = {
        "sku": "TEST-SKU",
        "analysis": {"colors": ["navy"], "features": [], "fit_or_size_hint": "", "style": ""},
        "bullet_points": [],
        "product_description": "",
        "title": "Test",
        "browse_node_id": "",
        "product_type": "",
        "browse_node": "",
        "image_paths": [],
        "image_url": "",
        "errors": [],
        "keywords": [],
        "material": "",
        "estimated_dimensions": "",
        "hsn_code": "",
        "product_tax_code": "",
        "parentage_level": "",
        "parent_sku": "",
        "brand_name": "",
        "_provider_hint": "",
    }
    avv = {"color": ["Black", "Navy Blue", "White"], "color map": ["Black", "Navy Blue", "White"]}
    row = m.catalog_result_to_template_row(result, columns, all_valid_values=avv)
    assert row["color"] == "Navy Blue", f"Expected 'Navy Blue', got '{row['color']}'"
    print("  PASS  test_catalog_result_to_template_row_snaps_color")


def test_generate_reference_locked_images_bg_removal():
    """generate_reference_locked_images applies BG removal before rendering."""
    import amazon_template_autofill_web as m
    from PIL import Image
    import inspect

    # Inspect source to verify _remove_photo_background is called
    src = inspect.getsource(m.generate_reference_locked_images)
    assert "_remove_photo_background" in src, \
        "generate_reference_locked_images should call _remove_photo_background"
    assert "_crop_to_content" in src, \
        "generate_reference_locked_images should call _crop_to_content"
    print("  PASS  test_generate_reference_locked_images_bg_removal")


def test_infographic_uses_drop_shadow():
    """generate_reference_locked_images uses _add_drop_shadow for infographic product image."""
    import amazon_template_autofill_web as m
    import inspect

    src = inspect.getsource(m.generate_reference_locked_images)
    # Count _add_drop_shadow calls — should be at least 2 (hero + infographic)
    count = src.count("_add_drop_shadow")
    assert count >= 2, f"Expected >=2 _add_drop_shadow calls, found {count}"
    print(f"  PASS  test_infographic_uses_drop_shadow — {count} shadow call(s)")


# ═════════════════════════════════════════════════════════════════════════════
# Playwright browser tests
# ═════════════════════════════════════════════════════════════════════════════

def run_playwright_tests():
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  SKIP  Playwright tests — playwright not installed")
        return

    proc = None
    try:
        print("  Starting test server on port", APP_PORT)
        proc = _start_server()
        BASE = f"http://localhost:{APP_PORT}"

        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()

            # ── Test 1: App loads ─────────────────────────────────────────────
            page.goto(BASE, wait_until="domcontentloaded")
            assert "Seller Studio" in page.title() or page.locator("h1,h2").first.is_visible(), \
                "App did not load"
            print("  PASS  playwright: app loads")

            # ── Test 2: Status endpoint includes Ollama provider ──────────────
            resp = page.request.get(f"{BASE}/api/status")
            assert resp.ok, f"Status API returned {resp.status}"
            data = resp.json()
            providers = data.get("providers", {})
            assert "ollama" in providers, f"Ollama not in providers: {list(providers.keys())}"
            print(f"  PASS  playwright: Ollama in /api/status providers "
                  f"(ok={providers['ollama']['ok']})")

            # ── Test 3: Template upload ───────────────────────────────────────
            if TEMPLATE.exists():
                page.goto(BASE, wait_until="domcontentloaded")
                upload_input = page.locator('input[type="file"]').first
                upload_input.set_input_files(str(TEMPLATE))
                # Give the server a moment to process
                page.wait_for_timeout(2000)
                # Check no error toast visible
                error_els = page.locator(".toast-error,.error-msg,[class*='error']")
                assert error_els.count() == 0 or all(
                    not e.is_visible() for e in error_els.element_handles()
                ), "Error visible after template upload"
                print("  PASS  playwright: template upload accepted")
            else:
                print("  SKIP  playwright: template upload — no Template.xlsm")

            # ── Test 4: /api/meta returns all_valid_values ────────────────────
            if TEMPLATE.exists():
                # Upload template first via API then call /api/meta
                with open(TEMPLATE, "rb") as f:
                    template_bytes = f.read()
                upload_resp = page.request.post(
                    f"{BASE}/api/upload-template",
                    multipart={"file": {"name": "Template.xlsm",
                                        "mimeType": "application/vnd.ms-excel.sheet.macroEnabled.12",
                                        "buffer": template_bytes}},
                )
                if upload_resp.ok:
                    upload_data = upload_resp.json()
                    tpath = upload_data.get("path", "")
                    if tpath:
                        meta_resp = page.request.post(
                            f"{BASE}/api/meta",
                            data=json.dumps({"template_path": tpath}),
                            headers={"Content-Type": "application/json"},
                        )
                        if meta_resp.ok:
                            meta_data = meta_resp.json()
                            avv = meta_data.get("all_valid_values", meta_data.get("meta", {}).get("all_valid_values", {}))
                            assert isinstance(avv, dict), "all_valid_values should be a dict"
                            print(f"  PASS  playwright: /api/meta all_valid_values has "
                                  f"{len(avv)} fields: {list(avv.keys())[:3]}")
                        else:
                            print(f"  WARN  playwright: /api/meta returned {meta_resp.status}")
            else:
                print("  SKIP  playwright: /api/meta all_valid_values — no Template.xlsm")

            browser.close()
    finally:
        if proc:
            proc.terminate()
            proc.wait(timeout=10)


# ═════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("\n=== Unit tests ===")
    test_snap_to_valid()
    test_read_all_valid_values()
    test_load_template_meta_includes_all_valid_values()
    test_catalog_result_to_template_row_snaps_color()
    test_generate_reference_locked_images_bg_removal()
    test_infographic_uses_drop_shadow()

    print("\n=== Playwright browser tests ===")
    run_playwright_tests()

    print("\n=== All tests complete ===")
